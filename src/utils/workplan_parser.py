"""
SAFMC Workplan Excel Parser

Parses SAFMC workplan Excel files and imports data into the database.
Handles the complex multi-sheet structure with amendments, milestones, and status tracking.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import re
import logging

logger = logging.getLogger(__name__)


class WorkplanParser:
    """Parse SAFMC workplan Excel files"""

    # Status code mappings from Excel
    STATUS_CODES = {
        'DOC': 'Document',
        'PH': 'Public Hearing',
        'A': 'Approved',
        'O/S': 'Outstanding',
        'AR': 'Approved for Review',
        '(AP)': 'AP Review',
        '(SSC)': 'SSC Review',
        'S': 'Scoping',
        'Snapper Grouper MSE': 'Management Strategy Evaluation',
        'Dolphin MSE': 'Management Strategy Evaluation',
    }

    # FMP/Committee mappings
    FMP_KEYWORDS = {
        'SG': 'Snapper Grouper',
        'CMP': 'Coastal Migratory Pelagics',
        'DW': 'Dolphin Wahoo',
        'Coral': 'Coral',
        'Shrimp': 'Shrimp',
        'Mackerel': 'Coastal Migratory Pelagics',
        'Spanish Mackerel': 'Coastal Migratory Pelagics',
        'King Mackerel': 'Coastal Migratory Pelagics',
    }

    def __init__(self, filepath: str):
        """Initialize parser with Excel file path"""
        self.filepath = filepath
        self.workbook = None
        self.worksheet = None

    def parse(self) -> Dict:
        """
        Parse the workplan Excel file

        Returns:
            dict: Parsed workplan data with items and milestones
        """
        try:
            self.workbook = openpyxl.load_workbook(self.filepath, data_only=True)

            # Try to find the main workplan sheet
            sheet_name = self._find_workplan_sheet()
            if not sheet_name:
                raise ValueError("Could not find workplan sheet in Excel file")

            self.worksheet = self.workbook[sheet_name]

            # Parse the structure
            header_row = self._find_header_row()
            quarter_columns = self._find_quarter_columns(header_row)

            # Parse amendment items
            items = self._parse_amendment_items(header_row, quarter_columns)

            logger.info(f"Parsed {len(items)} workplan items")

            return {
                'items': items,
                'quarter_columns': quarter_columns,
                'parse_date': datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Error parsing workplan: {e}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()

    def _find_workplan_sheet(self) -> Optional[str]:
        """Find the main workplan sheet"""
        # Common sheet names - updated for new format
        candidates = ['WorkPlan', 'Workplan', 'WORKPLAN', 'Sheet1', 'FMP Projects']

        for sheet_name in self.workbook.sheetnames:
            if any(candidate.lower() in sheet_name.lower() for candidate in candidates):
                return sheet_name

        # Default to first sheet if no match
        return self.workbook.sheetnames[0] if self.workbook.sheetnames else None

    def _find_header_row(self) -> int:
        """Find the row containing amendment headers (Amend #, Amendment, SAFMC Lead, etc.)"""
        for row_idx in range(1, 20):  # Search first 20 rows
            row = self.worksheet[row_idx]
            row_text = ' '.join([str(cell.value or '') for cell in row])

            # Look for characteristic headers - updated for new format
            if any(keyword in row_text for keyword in ['Amendment #', 'Topic', 'SAFMC Lead', 'SERO Priority']):
                return row_idx

        return 1  # Default to first row

    def _find_quarter_columns(self, header_row: int) -> Dict[str, int]:
        """
        Find columns for each quarter - handles both date formats:
        1. Old format: 'Dec-23', 'Mar-24', etc.
        2. New format: datetime objects like 2025-12-01, 2026-03-02

        Returns:
            dict: {quarter_label: column_index}
        """
        quarters = {}
        row = self.worksheet[header_row]

        for col_idx, cell in enumerate(row, start=1):
            cell_value = cell.value

            # Handle datetime objects (new format)
            if isinstance(cell_value, datetime):
                # Format as "Dec-25" style
                month_abbr = cell_value.strftime('%b')
                year_short = cell_value.strftime('%y')
                quarter_label = f"{month_abbr}-{year_short}"
                quarters[quarter_label] = col_idx
            # Handle string format (old format)
            elif isinstance(cell_value, str):
                quarter_pattern = r'(Dec|Mar|Jun|Sep)-(\d{2})'
                match = re.search(quarter_pattern, cell_value)
                if match:
                    quarters[cell_value.strip()] = col_idx

        return quarters

    def _parse_amendment_items(self, header_row: int, quarter_columns: Dict[str, int]) -> List[Dict]:
        """Parse all amendment items from the workplan"""
        items = []

        # Find column indices
        cols = self._find_column_indices(header_row)

        # Start parsing from row after header
        for row_idx in range(header_row + 1, self.worksheet.max_row + 1):
            row = self.worksheet[row_idx]

            # Get amendment number
            amend_num = self._get_cell_value(row, cols.get('amendment_number'))

            # Skip if no amendment number or if it's a subtotal/header row
            # Also skip if it's JUST a category marker (UNDERWAY/PLANNED without amendment number)
            if not amend_num:
                continue

            amend_num_str = str(amend_num).upper().strip()

            # Skip category-only rows or subtotal rows
            if amend_num_str in ['UNDERWAY', 'PLANNED', 'OTHER'] or \
               'SUBTOTAL' in amend_num_str or 'WORKLOAD' in amend_num_str:
                continue

            # Get amendment name
            amend_name = self._get_cell_value(row, cols.get('amendment_name'))
            if not amend_name:
                continue

            # Determine category
            category = self._determine_category(row_idx, header_row)

            # Get SAFMC lead
            safmc_lead = self._get_cell_value(row, cols.get('safmc_lead'))

            # Determine FMP
            fmp = self._determine_fmp(amend_num, amend_name)

            # Parse milestones for each quarter
            milestones = []
            for quarter_label, col_idx in quarter_columns.items():
                status = self._get_cell_value(row, col_idx)
                if status:
                    milestone = {
                        'quarter': quarter_label,
                        'quarter_date': self._parse_quarter_date(quarter_label),
                        'status_code': str(status).strip(),
                        'status_description': self.STATUS_CODES.get(str(status).strip(), str(status)),
                        'workload_value': self._extract_workload(status),
                        'is_complete': self._is_status_complete(status),
                        'is_pending': self._is_status_pending(status),
                    }
                    milestones.append(milestone)

            # Create item
            item = {
                'amendment_number': str(amend_num).strip(),
                'amendment_name': str(amend_name).strip(),
                'category': category,
                'safmc_lead': str(safmc_lead).strip() if safmc_lead else None,
                'fmp': fmp,
                'workload_points': sum(m['workload_value'] for m in milestones if m['workload_value']),
                'has_statutory_deadline': 'statutory deadline' in str(amend_name).lower(),
                'is_stock_assessment': 'assessment' in str(amend_name).lower() or 'stock' in str(amend_name).lower(),
                'milestones': milestones,
                'row_number': row_idx
            }

            items.append(item)

        return items

    def _find_column_indices(self, header_row: int) -> Dict[str, int]:
        """Find column indices for key fields"""
        cols = {}
        row = self.worksheet[header_row]

        for col_idx, cell in enumerate(row, start=1):
            cell_value = str(cell.value or '').lower().strip()

            # Support multiple formats:
            # 2025+: "Amendment #" and "Topic"
            # 2023: "Amend #" and "Amendment"

            # Check for amendment number column (various formats)
            if ('amendment' in cell_value or 'amend' in cell_value) and '#' in cell_value:
                cols['amendment_number'] = col_idx
            # Check for "Topic" column (2025+ format)
            elif cell_value == 'topic':
                cols['amendment_name'] = col_idx
            # Check for "Amendment" without # (2023 format)
            elif cell_value == 'amendment' and '#' not in cell_value:
                cols['amendment_name'] = col_idx
            # Check for "SAFMC Lead" column
            elif 'safmc' in cell_value and 'lead' in cell_value:
                cols['safmc_lead'] = col_idx

        return cols

    def _get_cell_value(self, row, col_idx):
        """Safely get cell value"""
        if not col_idx or col_idx > len(row):
            return None
        return row[col_idx - 1].value

    def _determine_category(self, row_idx: int, header_row: int) -> str:
        """Determine if amendment is underway, planned, or other"""
        # Look backwards for section headers
        for check_row in range(row_idx - 1, header_row, -1):
            first_cell = self.worksheet.cell(check_row, 1).value
            if first_cell:
                cell_text = str(first_cell).upper()
                if 'UNDERWAY' in cell_text:
                    return 'underway'
                elif 'PLANNED' in cell_text:
                    return 'planned'
                elif 'OTHER' in cell_text:
                    return 'other'

        return 'other'

    def _determine_fmp(self, amend_num: str, amend_name: str) -> str:
        """Determine FMP from amendment number or name"""
        combined = f"{amend_num} {amend_name}".upper()

        for keyword, fmp_name in self.FMP_KEYWORDS.items():
            if keyword.upper() in combined:
                return fmp_name

        return 'Other'

    def _parse_quarter_date(self, quarter_label: str) -> Optional[str]:
        """
        Parse quarter label to date
        e.g., 'Dec-23' -> '2023-12-01'
        """
        try:
            month_map = {'Dec': 12, 'Mar': 3, 'Jun': 6, 'Sep': 9}
            match = re.search(r'(Dec|Mar|Jun|Sep)-(\d{2})', quarter_label)
            if match:
                month_str, year_str = match.groups()
                month = month_map.get(month_str)
                year = 2000 + int(year_str)
                return f"{year}-{month:02d}-01"
        except:
            pass
        return None

    def _extract_workload(self, status) -> float:
        """Extract workload value from status (could be '1', '0.5', '2', etc.)"""
        if not status:
            return 0.0

        # Try to extract numeric value
        try:
            status_str = str(status).strip()
            # Look for numbers (including decimals)
            match = re.search(r'\d+\.?\d*', status_str)
            if match:
                return float(match.group())
        except:
            pass

        # Default workload based on status
        if status in ['DOC', 'PH', 'A']:
            return 1.0
        elif status in ['O/S', 'AR']:
            return 0.5

        return 0.0

    def _is_status_complete(self, status) -> bool:
        """Determine if status indicates completion"""
        if not status:
            return False
        status_str = str(status).upper()
        return 'A' == status_str or 'APPROVE' in status_str

    def _is_status_pending(self, status) -> bool:
        """Determine if status is pending/in progress"""
        if not status:
            return False
        status_str = str(status).upper()
        pending_indicators = ['O/S', 'OUTSTANDING', 'PENDING', '(', ')']
        return any(indicator in status_str for indicator in pending_indicators)


def parse_workplan_file(filepath: str) -> Dict:
    """
    Convenience function to parse a workplan file

    Args:
        filepath: Path to Excel file

    Returns:
        dict: Parsed workplan data
    """
    parser = WorkplanParser(filepath)
    return parser.parse()
