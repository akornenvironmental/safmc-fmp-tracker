"""
Workplan Excel Parser
Parses SAFMC workplan Excel files and extracts amendment data with timelines
"""

import logging
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class WorkplanParser:
    """Parser for SAFMC workplan Excel files"""

    # Milestone code mappings
    MILESTONE_CODES = {
        'AR': 'Assessment Report',
        'S': 'Scoping',
        'DOC': 'Document Review',
        'PH': 'Public Hearing',
        'A': 'Approval',
        'SUBMIT': 'Submitted',
        'IMPL': 'Implementation'
    }

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def parse_file(self, file_path: str) -> Dict:
        """
        Parse a workplan Excel file and return structured data

        Returns:
            {
                'metadata': {...},
                'amendments': [...],
                'timeline_headers': [...],
                'errors': [...]
            }
        """
        try:
            logger.info(f"Parsing workplan file: {file_path}")

            self.workbook = load_workbook(file_path, data_only=True)

            # Find the main workplan sheet
            workplan_sheet = self._find_workplan_sheet()
            if not workplan_sheet:
                raise ValueError("Could not find workplan sheet in Excel file")

            self.worksheet = self.workbook[workplan_sheet]

            # Extract metadata
            metadata = self._extract_metadata(file_path)

            # Extract column headers (timeline dates)
            timeline_headers = self._extract_timeline_headers()

            # Extract amendments
            amendments = self._extract_amendments(timeline_headers)

            result = {
                'metadata': metadata,
                'amendments': amendments,
                'timeline_headers': timeline_headers,
                'errors': []
            }

            logger.info(f"Successfully parsed {len(amendments)} amendments from workplan")

            return result

        except Exception as e:
            logger.error(f"Error parsing workplan file: {e}")
            import traceback
            traceback.print_exc()
            return {
                'metadata': {},
                'amendments': [],
                'timeline_headers': [],
                'errors': [str(e)]
            }

    def _find_workplan_sheet(self) -> Optional[str]:
        """Find the main workplan sheet"""
        # Look for sheets with names like "2025 - 2027 WorkPlan", "WorkPlan", etc.
        for sheet_name in self.workbook.sheetnames:
            if 'workplan' in sheet_name.lower() and 'how' not in sheet_name.lower():
                return sheet_name
        return None

    def _extract_metadata(self, file_path: str) -> Dict:
        """Extract metadata from workplan file"""
        metadata = {
            'file_name': file_path.split('/')[-1],
            'parsed_at': datetime.utcnow().isoformat()
        }

        # Try to extract quarter and year from filename
        # e.g., "fc2_a1_safmc_workplanq3_202509-xlsx.xlsx"
        filename = metadata['file_name']

        # Extract quarter
        quarter_match = re.search(r'q(\d)', filename, re.IGNORECASE)
        if quarter_match:
            metadata['quarter'] = f"Q{quarter_match.group(1)}"

        # Extract year
        year_match = re.search(r'20(\d{2})', filename)
        if year_match:
            metadata['fiscal_year'] = int(f"20{year_match.group(1)}")

        # Try to get title from sheet (Row 1)
        title_cell = self.worksheet.cell(row=1, column=2).value
        if title_cell:
            metadata['title'] = str(title_cell)

        return metadata

    def _extract_timeline_headers(self) -> List[Dict]:
        """
        Extract timeline column headers (dates)

        Row 3 contains the headers, starting from column 6 onwards
        """
        timeline_headers = []

        # Row 3 = headers, columns 6+ are timeline dates
        for col in range(6, self.worksheet.max_column + 1):
            header_val = self.worksheet.cell(row=3, column=col).value

            if header_val and isinstance(header_val, datetime):
                timeline_headers.append({
                    'column': col,
                    'date': header_val,
                    'meeting': header_val.strftime('%b %Y')  # e.g., "Dec 2025"
                })

        logger.info(f"Found {len(timeline_headers)} timeline columns")
        return timeline_headers

    def _extract_amendments(self, timeline_headers: List[Dict]) -> List[Dict]:
        """
        Extract amendment data from the worksheet

        Rows 4+ contain amendment data
        """
        amendments = []
        current_status = None

        for row_num in range(4, self.worksheet.max_row + 1):
            # Column 1 = Status section marker (UNDERWAY, PLANNED, etc.)
            status_val = self.worksheet.cell(row=row_num, column=1).value
            if status_val and isinstance(status_val, str):
                status_upper = status_val.strip().upper()
                if status_upper in ['UNDERWAY', 'PLANNED', 'COMPLETED', 'DEFERRED']:
                    current_status = status_upper
                    continue

            # Column 2 = Amendment ID
            amendment_id = self.worksheet.cell(row=row_num, column=2).value
            # Column 3 = Topic
            topic = self.worksheet.cell(row=row_num, column=3).value
            # Column 4 = Lead Staff
            lead_staff = self.worksheet.cell(row=row_num, column=4).value
            # Column 5 = SERO Priority
            sero_priority = self.worksheet.cell(row=row_num, column=5).value

            # Skip if no amendment ID or topic
            if not amendment_id or not topic:
                continue

            # Skip subtotal rows
            if 'SUBTOTAL' in str(topic).upper():
                continue

            # Extract milestones from timeline columns
            milestones = self._extract_milestones(row_num, timeline_headers)

            amendment_data = {
                'amendment_id': str(amendment_id).strip(),
                'topic': str(topic).strip(),
                'status': current_status or 'UNKNOWN',
                'lead_staff': str(lead_staff).strip() if lead_staff else None,
                'sero_priority': str(sero_priority).strip() if sero_priority else None,
                'milestones': milestones
            }

            amendments.append(amendment_data)

        return amendments

    def _extract_milestones(self, row_num: int, timeline_headers: List[Dict]) -> List[Dict]:
        """Extract milestone codes from timeline columns for a specific row"""
        milestones = []

        for timeline in timeline_headers:
            col = timeline['column']
            cell_val = self.worksheet.cell(row=row_num, column=col).value

            if not cell_val:
                continue

            # Cell may contain milestone codes like "S", "DOC", "PH/A", etc.
            cell_str = str(cell_val).strip()

            if not cell_str or cell_str.replace('.', '').isdigit():
                # Skip numeric values (workload estimates)
                continue

            # Parse milestone codes (may be combined like "PH/A")
            milestone_codes = self._parse_milestone_codes(cell_str)

            for code in milestone_codes:
                milestones.append({
                    'type': code,
                    'scheduled_date': timeline['date'],
                    'scheduled_meeting': timeline['meeting']
                })

        return milestones

    def _parse_milestone_codes(self, cell_value: str) -> List[str]:
        """
        Parse milestone codes from cell value

        Examples:
        - "S" -> ["S"]
        - "PH/A" -> ["PH", "A"]
        - "DOC" -> ["DOC"]
        """
        codes = []

        # Split by common separators
        parts = re.split(r'[/,\s]+', cell_value.strip())

        for part in parts:
            part_upper = part.strip().upper()

            # Check if it's a known milestone code
            if part_upper in self.MILESTONE_CODES:
                codes.append(part_upper)
            # Check for variations
            elif part_upper.startswith('PH'):
                codes.append('PH')
            elif part_upper.startswith('DOC'):
                codes.append('DOC')

        return codes


def parse_workplan_file(file_path: str) -> Dict:
    """
    Convenience function to parse a workplan file

    Usage:
        from src.scrapers.workplan_parser import parse_workplan_file
        result = parse_workplan_file('/path/to/workplan.xlsx')
    """
    parser = WorkplanParser()
    return parser.parse_file(file_path)
